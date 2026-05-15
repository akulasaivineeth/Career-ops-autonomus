"""
Dashboard data layer — reads SQLite + Excel and returns structured dicts
for the Jinja2 template.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

_DB_PATH = Path(__file__).resolve().parents[1] / "db" / "careerops.db"
_APPS_XLSX = Path(__file__).resolve().parents[1] / "data" / "applications.xlsx"
_PIPELINE_XLSX = Path(__file__).resolve().parents[1] / "data" / "pipeline.xlsx"
_REPORTS_DIR = Path(__file__).resolve().parents[1] / "reports"
_AUDIT_DIR = Path(__file__).resolve().parents[1] / "audit"


# ── helpers ─────────────────────────────────────────────────────────────────


def _conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _age(ts_str: str | None) -> str:
    if not ts_str:
        return "—"
    try:
        dt = datetime.fromisoformat(ts_str)
        delta = datetime.now() - dt
        if delta.total_seconds() < 60:
            return "just now"
        if delta.total_seconds() < 3600:
            return f"{int(delta.total_seconds() // 60)}m ago"
        if delta.days < 1:
            return f"{int(delta.total_seconds() // 3600)}h ago"
        return f"{delta.days}d ago"
    except Exception:
        return ts_str[:10] if ts_str else "—"


STATUS_STYLE: dict[str, tuple[str, str]] = {
    # (badge bg colour, text colour)
    "submitted": ("bg-green-100 text-green-800", "✅"),
    "apply_failed": ("bg-red-100 text-red-800", "❌"),
    "awaiting_approval": ("bg-yellow-100 text-yellow-800", "⏳"),
    "rejected_auto": ("bg-gray-100 text-gray-600", "⛔"),
    "queued": ("bg-blue-100 text-blue-800", "📋"),
    "scoring": ("bg-orange-100 text-orange-800", "🔍"),
    "tailoring": ("bg-purple-100 text-purple-800", "✍️"),
    "applying": ("bg-indigo-100 text-indigo-800", "🤖"),
    "paused": ("bg-gray-100 text-gray-500", "⏸"),
}


def _badge(status: str) -> dict[str, str]:
    cls, icon = STATUS_STYLE.get(status, ("bg-gray-100 text-gray-600", "•"))
    return {"cls": cls, "icon": icon, "label": status.replace("_", " ")}


# ── data fetchers ─────────────────────────────────────────────────────────────


def get_stats() -> dict[str, Any]:
    """Top-level KPIs."""
    if not _DB_PATH.exists():
        return {"total": 0, "submitted": 0, "pending": 0, "rejected": 0, "response_rate": 0}

    conn = _conn()
    rows = conn.execute("SELECT status, COUNT(*) as n FROM runs GROUP BY status").fetchall()
    conn.close()

    by_status: dict[str, int] = {r["status"]: r["n"] for r in rows}
    total = sum(by_status.values())
    submitted = by_status.get("submitted", 0)
    pending = by_status.get("awaiting_approval", 0)
    rejected = by_status.get("rejected_auto", 0) + by_status.get("apply_failed", 0)
    rate = round(submitted / total * 100, 1) if total else 0

    return {
        "total": total,
        "submitted": submitted,
        "pending": pending,
        "rejected": rejected,
        "response_rate": rate,
        "by_status": dict(by_status),
    }


def get_recent_runs(limit: int = 50) -> list[dict[str, Any]]:
    """Most recent apply runs with status badge."""
    if not _DB_PATH.exists():
        return []

    conn = _conn()
    rows = conn.execute(
        "SELECT run_id, jd_id, jd_url, track, score, ats_family, status, created_at, updated_at "
        "FROM runs ORDER BY created_at DESC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()

    result = []
    for r in rows:
        result.append(
            {
                "run_id": r["run_id"],
                "jd_id": r["jd_id"],
                "jd_url": r["jd_url"] or "",
                "track": r["track"] or "—",
                "score": f"{r['score']:.1f}" if r["score"] else "—",
                "ats": r["ats_family"] or "—",
                "status": r["status"] or "unknown",
                "badge": _badge(r["status"] or ""),
                "age": _age(r["created_at"]),
                "updated": _age(r["updated_at"]),
            }
        )
    return result


def get_pending_approvals() -> list[dict[str, Any]]:
    """Gates that are waiting for a human decision."""
    if not _DB_PATH.exists():
        return []

    conn = _conn()
    rows = conn.execute(
        "SELECT run_id, gate, sent_at, whatsapp_message_id "
        "FROM approvals WHERE decision IS NULL AND sent_at IS NOT NULL "
        "ORDER BY sent_at DESC LIMIT 20"
    ).fetchall()
    conn.close()

    return [
        {
            "run_id": r["run_id"],
            "gate": r["gate"],
            "sent": _age(r["sent_at"]),
            "msg_id": r["whatsapp_message_id"] or "—",
        }
        for r in rows
    ]


def get_pipeline(limit: int = 100) -> list[dict[str, Any]]:
    """Recent JDs from pipeline.xlsx / SQLite that haven't been applied to."""
    if not _PIPELINE_XLSX.exists():
        return []

    try:
        import openpyxl

        wb = openpyxl.load_workbook(_PIPELINE_XLSX, read_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hdr = {h: i for i, h in enumerate(headers) if h}

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(
                {
                    "jd_id": str(row[hdr.get("jd_id", 0)] or ""),
                    "company": str(row[hdr.get("company", 1)] or ""),
                    "role": str(row[hdr.get("role_title", 2)] or ""),
                    "source": str(row[hdr.get("source", 3)] or ""),
                    "status": str(row[hdr.get("status", 4)] or "queued"),
                    "badge": _badge(str(row[hdr.get("status", 4)] or "queued")),
                    "date": str(row[hdr.get("date_added", 5)] or ""),
                    "url": str(row[hdr.get("jd_url", 6)] or ""),
                }
            )
        wb.close()
        return rows[:limit]
    except Exception:
        return []


def get_score_chart_data() -> dict[str, Any]:
    """Score distribution for Chart.js bar chart."""
    if not _DB_PATH.exists():
        return {"labels": [], "data": [], "colors": []}

    conn = _conn()
    rows = conn.execute("SELECT score FROM runs WHERE score IS NOT NULL").fetchall()
    conn.close()

    buckets: dict[str, int] = {
        "1.0–1.9": 0,
        "2.0–2.9": 0,
        "3.0–3.9": 0,
        "4.0–4.4": 0,
        "4.5–5.0": 0,
    }
    for r in rows:
        s = float(r["score"])
        if s < 2:
            buckets["1.0–1.9"] += 1
        elif s < 3:
            buckets["2.0–2.9"] += 1
        elif s < 4:
            buckets["3.0–3.9"] += 1
        elif s < 4.5:
            buckets["4.0–4.4"] += 1
        else:
            buckets["4.5–5.0"] += 1

    colours = ["#ef4444", "#f97316", "#facc15", "#86efac", "#22c55e"]
    return {
        "labels": list(buckets.keys()),
        "data": list(buckets.values()),
        "colors": colours,
    }


def get_ats_chart_data() -> dict[str, Any]:
    """Submissions per ATS family for Chart.js doughnut."""
    if not _DB_PATH.exists():
        return {"labels": [], "data": [], "colors": []}

    conn = _conn()
    rows = conn.execute(
        "SELECT ats_family, COUNT(*) as n FROM runs WHERE status='submitted' GROUP BY ats_family"
    ).fetchall()
    conn.close()

    labels = [r["ats_family"] or "unknown" for r in rows]
    data = [r["n"] for r in rows]
    palette = [
        "#6366f1",
        "#8b5cf6",
        "#ec4899",
        "#14b8a6",
        "#f59e0b",
        "#10b981",
        "#3b82f6",
        "#f43f5e",
    ]
    colors = [palette[i % len(palette)] for i in range(len(labels))]
    return {"labels": labels, "data": data, "colors": colors}


def get_reports(limit: int = 20) -> list[dict[str, Any]]:
    """Recent score report markdown files."""
    if not _REPORTS_DIR.exists():
        return []

    files = sorted(_REPORTS_DIR.glob("*.md"), key=lambda p: p.stat().st_mtime, reverse=True)
    result = []
    for f in files[:limit]:
        text = f.read_text(encoding="utf-8")
        # Extract score line
        score = "—"
        for line in text.splitlines():
            if "**score:**" in line.lower():
                score = line.split(":")[-1].strip()
                break
        result.append(
            {
                "name": f.stem,
                "score": score,
                "age": _age(datetime.fromtimestamp(f.stat().st_mtime).isoformat()),
                "path": str(f),
            }
        )
    return result


def get_week_activity() -> dict[str, Any]:
    """Daily apply counts for the last 7 days — sparkline data."""
    if not _DB_PATH.exists():
        labels = [(datetime.now() - timedelta(days=i)).strftime("%a") for i in range(6, -1, -1)]
        return {"labels": labels, "data": [0] * 7}

    conn = _conn()
    labels, data = [], []
    for i in range(6, -1, -1):
        day = datetime.now() - timedelta(days=i)
        day_str = day.strftime("%Y-%m-%d")
        count = conn.execute(
            "SELECT COUNT(*) FROM runs WHERE created_at LIKE ?", (f"{day_str}%",)
        ).fetchone()[0]
        labels.append(day.strftime("%a"))
        data.append(count)
    conn.close()
    return {"labels": labels, "data": data}
