"""
Excel-first application tracker (BUILD.md extension, user request).

Two workbooks:
  data/applications.xlsx  — one row per submitted/tracked application
  data/pipeline.xlsx      — queued JDs awaiting scoring and apply

Both are kept in sync with SQLite (SQLite is the operational truth;
Excel is the human-readable view you open in Numbers/Excel).

Columns match the data contract (DATA_CONTRACT.md) plus the new
apply_status, apply_run_id, track, gate_log, screenshot_dir columns
added by Task 1.9 / migrate-applications-md.mjs.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import structlog
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = structlog.get_logger(__name__)

_DATA_DIR = Path(__file__).resolve().parents[2] / "data"
_APPS_XLSX = _DATA_DIR / "applications.xlsx"
_PIPELINE_XLSX = _DATA_DIR / "pipeline.xlsx"
_DB_PATH = Path(__file__).resolve().parents[2] / "db" / "careerops.db"

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

APPS_COLUMNS: list[tuple[str, int]] = [
    # (header, column width)
    ("run_id", 18),
    ("date", 13),
    ("company", 22),
    ("role_title", 35),
    ("jd_url", 55),
    ("ats_family", 14),
    ("track", 7),
    ("score", 8),
    ("apply_status", 18),
    ("gate_log", 30),
    ("screenshot_dir", 40),
    ("notes", 45),
]

PIPELINE_COLUMNS: list[tuple[str, int]] = [
    ("jd_id", 16),
    ("date_added", 13),
    ("company", 22),
    ("role_title", 35),
    ("source", 14),
    ("status", 18),
    ("jd_url", 55),
    ("notes", 45),
]

# Status → background colour (ARGB)
_STATUS_COLOURS: dict[str, str] = {
    "submitted": "FF90EE90",  # light green
    "apply_failed": "FFFFCCCB",  # light red
    "awaiting_approval": "FFFFFF99",  # light yellow
    "rejected_auto": "FFD3D3D3",  # light grey
    "queued": "FFE0FFFF",  # light cyan
    "scoring": "FFFFDAB9",  # peach
    "tailoring": "FFFFE4B5",  # moccasin
    "applying": "FFB0E0E6",  # powder blue
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _ensure_data_dir() -> None:
    _DATA_DIR.mkdir(parents=True, exist_ok=True)


def _header_style() -> Font:
    return Font(bold=True, color="FFFFFFFF", size=11)


def _header_fill() -> PatternFill:
    return PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")


def _apply_row_fill(ws: Any, row: int, status: str) -> None:
    colour = _STATUS_COLOURS.get(status)
    if not colour:
        return
    fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
    for cell in ws[row]:
        cell.fill = fill


def _create_sheet(wb: openpyxl.Workbook, title: str, columns: list[tuple[str, int]]) -> Any:
    ws = wb.active if wb.active and wb.active.title == "Sheet" else wb.create_sheet(title)
    ws.title = title
    ws.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _header_style()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return ws


def _header_map(ws: Any) -> dict[str, int]:
    """Return {column_name: column_index} from header row."""
    return {
        str(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value
    }


# ---------------------------------------------------------------------------
# Applications workbook
# ---------------------------------------------------------------------------


def _load_or_create_apps_wb() -> tuple[openpyxl.Workbook, Any]:
    if _APPS_XLSX.exists():
        wb = openpyxl.load_workbook(_APPS_XLSX)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = _create_sheet(wb, "Applications", APPS_COLUMNS)
    return wb, ws


def upsert_application(
    *,
    run_id: str,
    company: str = "",
    role_title: str = "",
    jd_url: str = "",
    ats_family: str = "",
    track: str = "",
    score: float | None = None,
    apply_status: str = "queued",
    gate_log: str = "",
    screenshot_dir: str = "",
    notes: str = "",
) -> None:
    """
    Insert or update one row in ``data/applications.xlsx``.

    Existing row is matched on ``run_id``.
    """
    _ensure_data_dir()
    wb, ws = _load_or_create_apps_wb()
    hdr = _header_map(ws)

    existing_row: int | None = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=hdr.get("run_id", 1)).value or "") == run_id:
            existing_row = row
            break

    target_row = existing_row or ws.max_row + 1

    def _set(col_name: str, value: Any) -> None:
        col = hdr.get(col_name)
        if col:
            ws.cell(row=target_row, column=col, value=value)

    _set("run_id", run_id)
    _set("date", datetime.now().strftime("%Y-%m-%d"))
    _set("company", company)
    _set("role_title", role_title)
    _set("jd_url", jd_url)
    _set("ats_family", ats_family)
    _set("track", track)
    _set("score", round(score, 2) if score is not None else "")
    _set("apply_status", apply_status)
    _set("gate_log", gate_log)
    _set("screenshot_dir", screenshot_dir)
    _set("notes", notes)

    _apply_row_fill(ws, target_row, apply_status)
    wb.save(_APPS_XLSX)
    log.info("tracker_upsert_app", run_id=run_id, status=apply_status)


def update_status(run_id: str, apply_status: str, *, notes: str = "") -> None:
    """Fast-path: update only apply_status (and optionally notes) for a run_id."""
    _ensure_data_dir()
    if not _APPS_XLSX.exists():
        log.warning("tracker_no_xlsx_to_update", run_id=run_id)
        return
    wb, ws = _load_or_create_apps_wb()
    hdr = _header_map(ws)
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=hdr.get("run_id", 1)).value or "") == run_id:
            status_col = hdr.get("apply_status")
            if status_col:
                ws.cell(row=row, column=status_col, value=apply_status)
            if notes and hdr.get("notes"):
                ws.cell(row=row, column=hdr["notes"], value=notes)
            _apply_row_fill(ws, row, apply_status)
            break
    wb.save(_APPS_XLSX)


# ---------------------------------------------------------------------------
# Pipeline workbook
# ---------------------------------------------------------------------------


def _load_or_create_pipeline_wb() -> tuple[openpyxl.Workbook, Any]:
    if _PIPELINE_XLSX.exists():
        wb = openpyxl.load_workbook(_PIPELINE_XLSX)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = _create_sheet(wb, "Pipeline", PIPELINE_COLUMNS)
    return wb, ws


def add_pipeline_jd(
    *,
    jd_id: str,
    company: str = "",
    role_title: str = "",
    source: str = "",
    jd_url: str = "",
    status: str = "queued",
    notes: str = "",
) -> None:
    """Append a new JD row to ``data/pipeline.xlsx`` (idempotent on jd_id)."""
    _ensure_data_dir()
    wb, ws = _load_or_create_pipeline_wb()
    hdr = _header_map(ws)

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=hdr.get("jd_id", 1)).value or "") == jd_id:
            log.debug("tracker_pipeline_already_exists", jd_id=jd_id)
            return

    target_row = ws.max_row + 1

    def _set(col_name: str, value: Any) -> None:
        col = hdr.get(col_name)
        if col:
            ws.cell(row=target_row, column=col, value=value)

    _set("jd_id", jd_id)
    _set("date_added", datetime.now().strftime("%Y-%m-%d"))
    _set("company", company)
    _set("role_title", role_title)
    _set("source", source)
    _set("status", status)
    _set("jd_url", jd_url)
    _set("notes", notes)
    _apply_row_fill(ws, target_row, status)

    wb.save(_PIPELINE_XLSX)
    log.info("tracker_pipeline_added", jd_id=jd_id)


def update_pipeline_status(jd_id: str, status: str) -> None:
    """Update the status of a pipeline row."""
    if not _PIPELINE_XLSX.exists():
        return
    wb, ws = _load_or_create_pipeline_wb()
    hdr = _header_map(ws)
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=hdr.get("jd_id", 1)).value or "") == jd_id:
            col = hdr.get("status")
            if col:
                ws.cell(row=row, column=col, value=status)
            _apply_row_fill(ws, row, status)
            break
    wb.save(_PIPELINE_XLSX)


# ---------------------------------------------------------------------------
# SQLite → Excel sync
# ---------------------------------------------------------------------------


def sync_from_db(db_path: Path | None = None) -> int:
    """
    Read all ``runs`` rows from SQLite and upsert them into ``applications.xlsx``.

    Returns the number of rows synced.
    """
    path = db_path or _DB_PATH
    if not path.exists():
        log.warning("tracker_sync_no_db")
        return 0

    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT run_id, jd_id, jd_url, track, score, ats_family, status, updated_at FROM runs"
    ).fetchall()
    conn.close()

    for row in rows:
        upsert_application(
            run_id=row["run_id"],
            jd_url=row["jd_url"] or "",
            ats_family=row["ats_family"] or "",
            track=row["track"] or "",
            score=float(row["score"]) if row["score"] is not None else None,
            apply_status=row["status"] or "",
        )

    log.info("tracker_synced", count=len(rows))
    return len(rows)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def _cli_summary() -> None:
    """Print a quick text table of today's status from applications.xlsx."""
    if not _APPS_XLSX.exists():
        print("No applications.xlsx yet — nothing tracked.")
        return

    wb = openpyxl.load_workbook(_APPS_XLSX, read_only=True)
    ws = wb.active
    hdr = _header_map(ws)
    status_col = hdr.get("apply_status", 0)
    company_col = hdr.get("company", 0)
    role_col = hdr.get("role_title", 0)
    date_col = hdr.get("date", 0)

    rows: list[tuple[str, str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(
            (
                str(row[date_col - 1] or ""),
                str(row[company_col - 1] or ""),
                str(row[role_col - 1] or ""),
                str(row[status_col - 1] or ""),
            )
        )
    wb.close()

    if not rows:
        print("No applications tracked yet.")
        return

    print(f"\n{'DATE':<13} {'COMPANY':<22} {'ROLE':<35} STATUS")
    print("-" * 90)
    for date, company, role, status in rows:
        print(f"{date:<13} {company:<22} {role:<35} {status}")
    print(f"\nTotal: {len(rows)}")


if __name__ == "__main__":
    _cli_summary()
