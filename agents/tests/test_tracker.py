"""Excel tracker tests (tracker.py)."""

from __future__ import annotations

from pathlib import Path

import pytest

from agents.tools import tracker as t_mod
from agents.tools.tracker import (
    add_pipeline_jd,
    sync_from_db,
    update_pipeline_status,
    update_status,
    upsert_application,
)


@pytest.fixture()
def workdir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(t_mod, "_DATA_DIR", tmp_path)
    monkeypatch.setattr(t_mod, "_APPS_XLSX", tmp_path / "applications.xlsx")
    monkeypatch.setattr(t_mod, "_PIPELINE_XLSX", tmp_path / "pipeline.xlsx")
    return tmp_path


def test_upsert_creates_file(workdir: Path) -> None:
    upsert_application(
        run_id="run-001",
        company="Acme Corp",
        role_title="ML Engineer",
        jd_url="https://example.com/job/1",
        apply_status="queued",
    )
    assert (workdir / "applications.xlsx").exists()


def test_upsert_idempotent(workdir: Path) -> None:
    for status in ("queued", "applying", "submitted"):
        upsert_application(run_id="run-002", company="Beta", apply_status=status)
    import openpyxl

    wb = openpyxl.load_workbook(workdir / "applications.xlsx")
    ws = wb.active
    rows = [
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert rows.count("run-002") == 1, "same run_id must not duplicate"


def test_update_status(workdir: Path) -> None:
    upsert_application(run_id="run-003", company="Gamma", apply_status="queued")
    update_status("run-003", "submitted")
    import openpyxl

    wb = openpyxl.load_workbook(workdir / "applications.xlsx")
    ws = wb.active
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=hdr["run_id"]).value == "run-003":
            assert ws.cell(row=row, column=hdr["apply_status"]).value == "submitted"
            break


def test_add_pipeline_jd(workdir: Path) -> None:
    add_pipeline_jd(
        jd_id="JD-001",
        company="Delta",
        role_title="DE",
        source="greenhouse",
        jd_url="https://x.com",
    )
    assert (workdir / "pipeline.xlsx").exists()


def test_add_pipeline_idempotent(workdir: Path) -> None:
    for _ in range(3):
        add_pipeline_jd(
            jd_id="JD-002",
            company="Epsilon",
            role_title="MLE",
            source="lever",
            jd_url="https://y.com",
        )
    import openpyxl

    wb = openpyxl.load_workbook(workdir / "pipeline.xlsx")
    ws = wb.active
    rows = [
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert rows.count("JD-002") == 1


def test_update_pipeline_status(workdir: Path) -> None:
    add_pipeline_jd(
        jd_id="JD-003", company="Zeta", role_title="DA", source="ashby", jd_url="https://z.com"
    )
    update_pipeline_status("JD-003", "applying")
    import openpyxl

    wb = openpyxl.load_workbook(workdir / "pipeline.xlsx")
    ws = wb.active
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=hdr["jd_id"]).value == "JD-003":
            assert ws.cell(row=row, column=hdr["status"]).value == "applying"


def test_sync_from_empty_db(workdir: Path, tmp_path: Path) -> None:
    db = tmp_path / "empty.db"
    from db.migrate import apply_migrations

    apply_migrations(db)
    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(t_mod, "_DB_PATH", db)
    count = sync_from_db(db)
    assert count == 0
    monkeypatch.undo()
